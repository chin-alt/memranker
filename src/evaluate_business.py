from __future__ import annotations

import argparse
import csv
import json
import logging
import math
import re
import time

from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from data import format_input_text, record_to_doc, write_jsonl
from modeling import DEFAULT_MODEL_NAME, load_scorer


logger = logging.getLogger(__name__)

DEFAULT_BUSINESS_INSTRUCTION = (
    "Given a user query, retrieve relevant documents that answer the query."
)
DOC_ID_SPLIT_RE = re.compile(r"[,，、;；\n\r\t]+")


@dataclass
class GroundTruthItem:
    query: str
    raw_page_ids: str = ""
    doc_ids: list[str] = field(default_factory=list)

    @property
    def doc_id_set(self) -> set[str]:
        return set(self.doc_ids)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Evaluate Qwen3 yes/no-logit reranker on a business recall dataset."
    )
    parser.add_argument("--gt_file", required=True, help="Excel/CSV file with query-doc ground truth.")
    parser.add_argument("--recall_file", required=True, help="JSON file with recalled docs per query.")
    parser.add_argument("--model_path", default=DEFAULT_MODEL_NAME, help="Base model or finetuned checkpoint.")
    parser.add_argument("--output_dir", default="outputs/business_eval")
    parser.add_argument("--instruction", default=DEFAULT_BUSINESS_INSTRUCTION)
    parser.add_argument("--gt_query_col", default="query")
    parser.add_argument("--gt_doc_id_col", default="PageId")
    parser.add_argument("--gt_sheet", default=None, help="Optional Excel sheet name. Defaults to active sheet.")
    parser.add_argument("--recall_id_key", default="id")
    parser.add_argument("--recall_text_key", default="text")
    parser.add_argument("--max_length", type=int, default=4096)
    parser.add_argument("--batch_size", type=int, default=4)
    parser.add_argument("--top_k_list", type=int, nargs="+", default=[1, 3, 5, 10])
    parser.add_argument("--attn_implementation", default=None)
    parser.add_argument("--bf16", action="store_true")
    parser.add_argument("--fp16", action="store_true")
    parser.add_argument("--mock", action="store_true", help="Use lexical mock scorer for smoke tests.")
    parser.add_argument(
        "--save_doc_text",
        action="store_true",
        help="Include full document text in predictions.jsonl for debugging.",
    )
    return parser.parse_args()


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def split_doc_ids(value: Any) -> list[str]:
    raw = clean_cell(value)
    if not raw:
        return []
    doc_ids: list[str] = []
    seen: set[str] = set()
    for part in DOC_ID_SPLIT_RE.split(raw):
        doc_id = clean_cell(part).strip(" \u3000")
        if not doc_id or doc_id in seen:
            continue
        seen.add(doc_id)
        doc_ids.append(doc_id)
    return doc_ids


def add_ground_truth_item(
    gt: dict[str, GroundTruthItem],
    query: str,
    raw_page_ids: str,
    doc_ids: list[str],
) -> None:
    item = gt.setdefault(query, GroundTruthItem(query=query))
    if raw_page_ids:
        item.raw_page_ids = raw_page_ids if not item.raw_page_ids else f"{item.raw_page_ids}，{raw_page_ids}"
    seen = set(item.doc_ids)
    for doc_id in doc_ids:
        if doc_id not in seen:
            item.doc_ids.append(doc_id)
            seen.add(doc_id)


def load_ground_truth(
    gt_file: str | Path,
    query_col: str,
    doc_id_col: str,
    sheet_name: str | None = None,
) -> dict[str, GroundTruthItem]:
    gt: dict[str, GroundTruthItem] = {}
    skipped = 0
    path = Path(gt_file)

    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames or []
            missing = [col for col in (query_col, doc_id_col) if col not in fieldnames]
            if missing:
                raise ValueError(f"Ground-truth file missing columns: {missing}. Found: {fieldnames}")
            for row in reader:
                query = clean_cell(row.get(query_col))
                raw_page_ids = clean_cell(row.get(doc_id_col))
                doc_ids = split_doc_ids(raw_page_ids)
                if not query or not doc_ids:
                    skipped += 1
                    continue
                add_ground_truth_item(gt, query, raw_page_ids, doc_ids)
    else:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Business Excel evaluation requires openpyxl. "
                "Install it with: pip install openpyxl"
            ) from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet {sheet_name!r} not found. Available sheets: {workbook.sheetnames}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError(f"Ground-truth file has no header row: {gt_file}")
        headers = [clean_cell(value) for value in header_row]
        missing = [col for col in (query_col, doc_id_col) if col not in headers]
        if missing:
            raise ValueError(f"Ground-truth file missing columns: {missing}. Found: {headers}")
        query_idx = headers.index(query_col)
        doc_id_idx = headers.index(doc_id_col)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            query = clean_cell(row[query_idx] if query_idx < len(row) else None)
            raw_page_ids = clean_cell(row[doc_id_idx] if doc_id_idx < len(row) else None)
            doc_ids = split_doc_ids(raw_page_ids)
            if not query or not doc_ids:
                skipped += 1
                continue
            add_ground_truth_item(gt, query, raw_page_ids, doc_ids)

    if skipped:
        logger.warning("Skipped %d ground-truth rows with empty query/doc id", skipped)
    logger.info("Loaded ground truth for %d queries from %s", len(gt), gt_file)
    return dict(gt)


def stringify_doc_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def coerce_recall_doc(
    item: Any,
    id_key: str,
    text_key: str,
) -> tuple[str, str] | None:
    if not isinstance(item, dict):
        return None

    doc_id = clean_cell(
        item.get(id_key)
        or item.get("doc_id")
        or item.get("page_id")
        or item.get("PageId")
        or item.get("id")
    )
    text = stringify_doc_text(item.get(text_key))
    if not text:
        text = record_to_doc(item)
    if not doc_id or not text:
        return None
    return doc_id, text


def load_recall_results(
    recall_file: str | Path,
    id_key: str,
    text_key: str,
) -> dict[str, list[dict[str, str]]]:
    with Path(recall_file).open("r", encoding="utf-8-sig") as f:
        data = json.load(f)

    recall: dict[str, list[dict[str, str]]] = defaultdict(list)
    skipped = 0

    if isinstance(data, dict):
        iterable = data.items()
    elif isinstance(data, list):
        rows = []
        for row in data:
            if not isinstance(row, dict):
                skipped += 1
                continue
            query = clean_cell(row.get("query") or row.get("q") or row.get("question"))
            docs = row.get("docs") or row.get("documents") or row.get("recall") or row.get("items")
            if query and isinstance(docs, list):
                rows.append((query, docs))
            elif query:
                rows.append((query, [row]))
            else:
                skipped += 1
        iterable = rows
    else:
        raise ValueError("Recall JSON must be either a dict or a list.")

    for query_raw, docs_raw in iterable:
        query = clean_cell(query_raw)
        if not query or not isinstance(docs_raw, list):
            skipped += 1
            continue
        for item in docs_raw:
            coerced = coerce_recall_doc(item, id_key=id_key, text_key=text_key)
            if coerced is None:
                skipped += 1
                continue
            doc_id, doc_text = coerced
            recall[query].append({"doc_id": doc_id, "doc": doc_text})

    if skipped:
        logger.warning("Skipped %d malformed recall rows/docs from %s", skipped, recall_file)
    logger.info("Loaded recall docs for %d queries from %s", len(recall), recall_file)
    return dict(recall)


def build_scoring_inputs(
    recall_results: dict[str, list[dict[str, str]]],
    ground_truth: dict[str, GroundTruthItem],
    instruction: str,
) -> tuple[list[str], list[dict[str, Any]], int]:
    input_texts: list[str] = []
    mapping: list[dict[str, Any]] = []
    skipped_queries = 0

    for query, docs in recall_results.items():
        if query not in ground_truth:
            skipped_queries += 1
            continue
        for idx, doc in enumerate(docs):
            doc_text = doc["doc"]
            input_texts.append(format_input_text(instruction, query, doc_text))
            mapping.append(
                {
                    "query": query,
                    "doc_id": doc["doc_id"],
                    "doc": doc_text,
                    "source_rank": idx + 1,
                }
            )

    return input_texts, mapping, skipped_queries


def attach_scores_and_ranks(
    mapping: list[dict[str, Any]],
    scores: list[float],
    ground_truth: dict[str, GroundTruthItem],
    save_doc_text: bool,
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row, score in zip(mapping, scores, strict=False):
        gt_item = ground_truth.get(row["query"])
        gt_ids = gt_item.doc_id_set if gt_item is not None else set()
        item = {
            "query": row["query"],
            "doc_id": row["doc_id"],
            "score": float(score),
            "source_rank": row["source_rank"],
            "is_relevant": row["doc_id"] in gt_ids,
        }
        if save_doc_text:
            item["doc"] = row["doc"]
        grouped[row["query"]].append(item)

    ranked: list[dict[str, Any]] = []
    for query in sorted(grouped):
        rows = sorted(
            grouped[query],
            key=lambda item: (-float(item["score"]), int(item["source_rank"])),
        )
        for rank, row in enumerate(rows, start=1):
            row["rank"] = rank
            ranked.append(row)
    return ranked


def join_ids(values: list[str]) -> str:
    return "，".join(values)


SUMMARY_COLUMNS = [
    "query",
    "PageId",
    "正确标签数量",
    "召回候选数量",
    "模型召回的ID",
    "命中ID",
    "漏召ID",
    "命中数量",
    "准确率",
    "推理时间(s)",
]


def compute_business_metrics(
    ranked_predictions: list[dict[str, Any]],
    ground_truth: dict[str, GroundTruthItem],
    top_k_list: list[int],
    seconds_per_example: float,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for pred in ranked_predictions:
        grouped[pred["query"]].append(pred)

    per_query_rows: list[dict[str, Any]] = []
    metrics: dict[str, Any] = {
        "num_gt_queries": len(ground_truth),
        "num_scored_queries": len(grouped),
        "num_scored_pairs": len(ranked_predictions),
    }
    accuracy_sum = 0.0
    total_hits_at_label_count = 0
    total_gt_docs = 0

    for query, gt_item in ground_truth.items():
        gt_doc_ids = gt_item.doc_id_set
        preds = sorted(grouped.get(query, []), key=lambda row: int(row.get("rank", 10**9)))
        label_count = len(gt_item.doc_ids)
        model_top_by_label_count = preds[:label_count]
        model_top_ids = [str(pred["doc_id"]) for pred in model_top_by_label_count]
        model_top_id_set = set(model_top_ids)
        hit_ids = [doc_id for doc_id in model_top_ids if doc_id in gt_doc_ids]
        missed_ids = [doc_id for doc_id in gt_item.doc_ids if doc_id not in model_top_id_set]
        hits_at_label_count = len(set(model_top_ids) & gt_doc_ids)
        accuracy_at_label_count = hits_at_label_count / label_count if label_count else 0.0
        estimated_query_time = len(preds) * seconds_per_example

        row: dict[str, Any] = {
            "query": query,
            "PageId": gt_item.raw_page_ids or join_ids(gt_item.doc_ids),
            "正确标签数量": label_count,
            "召回候选数量": len(preds),
            "模型召回的ID": join_ids(model_top_ids),
            "命中ID": join_ids(hit_ids),
            "漏召ID": join_ids(missed_ids),
            "命中数量": hits_at_label_count,
            "准确率": accuracy_at_label_count,
            "推理时间(s)": estimated_query_time,
        }
        first_hit_rank = next((int(pred["rank"]) for pred in preds if pred["doc_id"] in gt_doc_ids), 0)
        row["MRR"] = 0.0 if first_hit_rank == 0 else 1.0 / first_hit_rank

        for top_k in top_k_list:
            top_preds = preds[:top_k]
            top_ids = [pred["doc_id"] for pred in top_preds]
            hits = len(set(top_ids) & gt_doc_ids)
            precision = hits / len(top_ids) if top_ids else 0.0
            recall = hits / len(gt_doc_ids) if gt_doc_ids else 0.0
            f1 = 0.0 if precision + recall == 0 else 2 * precision * recall / (precision + recall)
            row[f"Precision@{top_k}"] = precision
            row[f"Recall@{top_k}"] = recall
            row[f"F1@{top_k}"] = f1
            row[f"HitRate@{top_k}"] = 1.0 if hits > 0 else 0.0

        per_query_rows.append(row)
        accuracy_sum += accuracy_at_label_count
        total_hits_at_label_count += hits_at_label_count
        total_gt_docs += label_count

    denom = max(1, len(per_query_rows))
    metrics["Accuracy@GTCount"] = accuracy_sum / denom
    metrics["MicroAccuracy@GTCount"] = total_hits_at_label_count / total_gt_docs if total_gt_docs else 0.0
    metrics["total_gt_docs"] = total_gt_docs
    metrics["total_hits_at_gt_count"] = total_hits_at_label_count
    metrics["MRR"] = sum(float(row["MRR"]) for row in per_query_rows) / denom
    for top_k in top_k_list:
        for name in ("Precision", "Recall", "F1", "HitRate"):
            key = f"{name}@{top_k}"
            metrics[key] = sum(float(row[key]) for row in per_query_rows) / denom
    return metrics, per_query_rows


def write_summary_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=SUMMARY_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_summary_xlsx(path: Path, rows: list[dict[str, Any]]) -> bool:
    try:
        from openpyxl import Workbook
    except ImportError:
        logger.warning("openpyxl is not installed; skipped xlsx summary output.")
        return False

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "business_eval"
    sheet.append(SUMMARY_COLUMNS)
    for row in rows:
        sheet.append([row.get(col, "") for col in SUMMARY_COLUMNS])
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(12, max_length + 2), 60)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return True


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s - %(message)s")
    args = parse_args()
    if args.bf16 and args.fp16:
        raise ValueError("--bf16 and --fp16 are mutually exclusive")

    ground_truth = load_ground_truth(
        args.gt_file,
        query_col=args.gt_query_col,
        doc_id_col=args.gt_doc_id_col,
        sheet_name=args.gt_sheet,
    )
    recall_results = load_recall_results(
        args.recall_file,
        id_key=args.recall_id_key,
        text_key=args.recall_text_key,
    )
    input_texts, mapping, skipped_queries = build_scoring_inputs(
        recall_results,
        ground_truth,
        instruction=args.instruction,
    )
    if skipped_queries:
        logger.warning("Skipped %d recall queries not found in ground truth", skipped_queries)
    if not input_texts:
        raise ValueError("No query-document pairs to score after matching recall data to ground truth.")

    scorer = load_scorer(
        args.model_path,
        max_length=args.max_length,
        bf16=args.bf16,
        fp16=args.fp16,
        mock=args.mock,
        attn_implementation=args.attn_implementation,
    )

    logger.info(
        "Scoring %d business query-document pairs with Qwen3 standard reranker prompt",
        len(input_texts),
    )
    start_time = time.perf_counter()
    scores = scorer.predict(input_texts, batch_size=args.batch_size)
    score_time = time.perf_counter() - start_time
    sec_per_example = score_time / max(1, len(input_texts))
    examples_per_sec = len(input_texts) / score_time if score_time > 0 else 0.0

    ranked_predictions = attach_scores_and_ranks(
        mapping,
        scores,
        ground_truth,
        save_doc_text=args.save_doc_text,
    )
    metrics, per_query = compute_business_metrics(
        ranked_predictions,
        ground_truth,
        args.top_k_list,
        seconds_per_example=sec_per_example,
    )
    metrics.update(
        {
            "score_time_seconds": float(score_time),
            "seconds_per_example": float(sec_per_example),
            "examples_per_second": float(examples_per_sec),
            "skipped_recall_queries_without_gt": skipped_queries,
        }
    )

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    write_jsonl(output_dir / "per_query_metrics.jsonl", per_query)
    write_jsonl(output_dir / "predictions.jsonl", ranked_predictions)
    write_summary_csv(output_dir / "business_eval.csv", per_query)
    wrote_xlsx = write_summary_xlsx(output_dir / "business_eval.xlsx", per_query)
    metrics["summary_csv"] = str(output_dir / "business_eval.csv")
    metrics["summary_xlsx"] = str(output_dir / "business_eval.xlsx") if wrote_xlsx else ""
    (output_dir / "metrics.json").write_text(
        json.dumps(metrics, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    logger.info("Wrote business evaluation outputs to %s", output_dir)
    print(json.dumps(metrics, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
